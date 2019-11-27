import tkinter as tk
import openpyxl
import copy
import os
import datetime

from tkinter import messagebox
from tkinter import ttk

# selectの表示は配列の上書きを利用して表示
# 記入した日付を配列に入れる処理をする

path_teamdata = os.path.abspath("excel/teamdata.xlsx")
path_db = os.path.abspath("db.xlsx")

#open flie teamdata
wb_teamdata = openpyxl.load_workbook(path_teamdata)
sheet_wb_t = wb_teamdata.sheetnames[0]
ws_teamdata = wb_teamdata[sheet_wb_t]

# 初期化
players_information = []
symptom_information = []

players_info_dict = {}
symptom_info_dict = {}

body_table_dict = {}
name_table_dict = {}
symptom_table_dict = {}

# open file db
wb_db = openpyxl.load_workbook(path_db)
ws_players = wb_db["players_info"]
ws_symptom = wb_db["symptom_info"]

ws_body_table = wb_db["body_table"]
ws_name_table = wb_db["name_table"]
ws_symptom_table = wb_db["symptom_table"]

# wsを返す
def return_ws_db(sheetname):
    # wb_db = openpyxl.load_workbook(path_db)
    ws = wb_db[sheetname]
    return ws

#1行目の値と列番号をlistに変換
for row in ws_teamdata.iter_rows(min_row=1):
    for cell in row:
        val = cell.value
        players_information.append(val)


# valueからcolumnのindexを取得する関数
def get_keys_from_value(d, val):
    keys = [k for k, v in d.items() if v == val]
    if keys:
        return keys[0]

def make_dict_first_col_value(sheet_name, dictname):
    for rows in sheet_name.iter_rows(min_row=1, max_row=1):
        for cell in rows:
            val = cell.value
            col_index = cell.col_idx
            dictname[col_index] = val

# idをkeyにして値がvalueになるdict
def id_value_key_dict(sheet_name, min=None, max=None):
    d = {}
    l = []
    for cols in sheet_name.iter_cols(min_row=min,max_row=max):
        a = []
        for cell in cols:
            val = cell.value
            a.append(val)
        c = copy.deepcopy(a)
        l.append(c)
    new_d = dict(zip(l[1], l[0]))
    return new_d

# idとvalueをdictにまとめる処理
def id_num_key_dict(sheet_name, min=None, max=None):
    d = {}
    l = []
    for cols in sheet_name.iter_cols(min_row=min,max_row=max):
        a = []
        for cell in cols:
            val = cell.value
            a.append(val)
        c = copy.deepcopy(a)
        l.append(c)
    new_d = dict(zip(l[0], l[1]))
    return new_d


def row_list(ws, min=None, max=None):
    list = []
    a = []
    for col in ws.iter_rows(min_row=min, max_row=max):
        for cell in col:
            a.append(cell.value)
        c = copy.deepcopy(a)
        a.clear()
        list.append(c)
    return list

def col_list(ws, min=None, max=None):
    list = []
    a = []
    for col in ws.iter_cols(min_row=min, max_row=max):
        for cell in col:
            a.append(cell.value)
        c = copy.deepcopy(a)
        a.clear()
        list.append(c)
    return list

def make_dict_colvalues(ws,list):
    i = 0
    a = []
    d = {}
    for col in ws.iter_cols(min_col=1, min_row=2):
        for cell in col:
            val = cell.value
            a.append(val)
        c = copy.deepcopy(a)
        d[list[i]] = c
        a.clear()
        i += 1
    return d


def change_page(page):
    page.tkraise()

def make_row_list(list, sheetname, m_row=None):
    for rows in sheetname.iter_rows(min_row=m_row):
        for cell in rows:
            list.append(cell.value)

def return_row_index(sheetname, id):
    d = {}
    a = 0
    for cols in sheetname.iter_cols(max_col=1):
        for cell in cols:
            a += 1
            d[cell.value] = a
    try:
        re = d[id]
        return re
    except KeyError as e:
        return None

def show_select_info(name, dict):
    name_id = dict[name]
    # symptom_infoシートの全情報をリスト化
    informations = row_list(ws_symptom)
    row_num = return_row_index(ws_symptom, name_id)
    select_info = informations[row_num - 1] #indexのため-1
    print(informations)

    verletzungsart_id = select_info[1]
    korperteil_id = select_info[2]

    # 各idを挿入
    info_frame.entry_name_id.delete(0, tk.END)
    info_frame.entry_name_id.insert(tk.END, name_id)
    info_frame.entry_verletzungsart_id.delete(0, tk.END)
    info_frame.entry_verletzungsart_id.insert(tk.END, verletzungsart_id)
    info_frame.entry_korperteil_id.delete(0, tk.END)
    info_frame.entry_korperteil_id.insert(tk.END, korperteil_id)

    # 各シートのidとvalueをdictにまとめる
    verletzungsart_info = id_num_key_dict(ws_symptom_table)
    korperteil_info = id_num_key_dict(ws_body_table)

    # idをkeyにしてvalueを取得
    verletzungsart = verletzungsart_info[verletzungsart_id]
    korperteil = korperteil_info[korperteil_id]

    select_info[1] = verletzungsart
    select_info[2] = korperteil

    # info_frame.entry_verletzungsart.delete(0, tk.END)
    # info_frame.entry_verletzungsart.insert(tk.END, verletzungsart)
    info_frame.verletzungsarts[0] = verletzungsart
    info_frame.combo_verletzungsart["value"] = info_frame.verletzungsarts
    info_frame.combo_verletzungsart.current(0)
    # info_frame.entry_korperteil.delete(0, tk.END)
    # info_frame.entry_korperteil.insert(tk.END, korperteil)
    info_frame.korperteils[0] = korperteil
    info_frame.combo_korperteil["value"] = info_frame.korperteils
    info_frame.combo_korperteil.current(0)

    info_frame.entry_verletzungsdatum.delete(0, tk.END)
    info_frame.entry_verletzungsdatum.insert(tk.END, select_info[3])
    info_frame.entry_der_geheiligte_tag.delete(0, tk.END)
    info_frame.entry_der_geheiligte_tag.insert(tk.END, select_info[4])
    info_frame.entry_wie.delete(0, tk.END)
    info_frame.entry_wie.insert(tk.END, select_info[5])
    info_frame.entry_verletzter_ort.delete(0, tk.END)
    info_frame.entry_verletzter_ort.insert(tk.END, select_info[6])
    info_frame.entry_diagnose.delete(0, tk.END)
    info_frame.entry_diagnose.insert(tk.END, select_info[7])
    info_frame.entry_sonstig.delete(0, tk.END)
    info_frame.entry_sonstig.insert(tk.END, select_info[8])


def show_alert():
    ret = messagebox.askyesno('確認', '登録しますか？')
    return ret

# 選手情報の登録 動的に取得したったけどできなかった。
def register_player(name, age, sex, position):
    name_text = name.get()
    age_text = int(age.get())
    sex_text = int(sex.get())
    position_text = position.get()

    ret = messagebox.askyesno("確認",
                            f"""{players_information[1]}:{name_text}\n/{players_information[2]}:{age_text}\n{players_information[3]}:{sex_text}\n{players_information[4]}:{position_text}\n
                            """)
    # 登録処理
    if ret == True:
        players_info_row = ws_players.max_row + 1
        ws_name_taple_row = ws_name_table.max_row + 1
        symptom_info_row = ws_symptom.max_row + 1

        get_keys_name = get_keys_from_value(players_info_dict, "Name_id")
        get_keys_sex = get_keys_from_value(players_info_dict, players_information[2])
        get_keys_age = get_keys_from_value(players_info_dict, players_information[3])
        get_keys_position = get_keys_from_value(players_info_dict, players_information[4])

        init_check = ws_players.cell(row=ws_players.max_row, column=get_keys_name).value
        if ws_players.max_row is not 1:
            info_previous_val = ws_players.cell(row=ws_players.max_row, column=get_keys_name).value
            table_previous_val = ws_name_table.cell(row=ws_name_table.max_row, column=1).value
            info_symptom_previous_val = ws_symptom.cell(row=ws_symptom.max_row, column=1).value

            ws_players.cell(row=players_info_row, column=get_keys_name).value = info_previous_val + 1
            ws_name_table.cell(row=players_info_row, column=1).value = table_previous_val + 1
            ws_symptom.cell(row=players_info_row, column=1).value = info_symptom_previous_val + 1
        else:
            ws_players.cell(row=players_info_row, column=get_keys_name).value = 1
            ws_name_table.cell(row=players_info_row, column=1).value = 1
            ws_symptom.cell(row=players_info_row, column=1).value = 1
        # players_info sheet
        ws_players.cell(row=players_info_row, column=get_keys_sex[0]).value = sex_text
        ws_players.cell(row=players_info_row, column=get_keys_age[0]).value = age_text
        ws_players.cell(row=players_info_row, column=get_keys_position[0]).value = position_text
        # name_table sheet
        ws_name_table.cell(row=ws_name_taple_row, column=2).value = name_text
        # symptom_table sheet
        ws_symptom_table.cell(row=ws_symptom_table_row, column=2).value

        wb_db.save(path_db)

        info_frame.tkraise()

# 選手の状態を登録
def register_symptom(*args):
    a = []
    # シートの値を取得
    symptom_i = return_ws_db("symptom_info")
    symptom_t = return_ws_db("symptom_table")
    name_t = return_ws_db("name_table")
    body_t = return_ws_db("body_table")

    # 取得した値を配列に変換
    for arg in args:
        arg_get = arg.get()
        a.append(arg_get)

    for i in range(3):
        id = int(a[i])
        a[i] = id

    day =  datetime.date.today()
    a.append(day)
    print(a)

    length = len(a)
    print(symptom_i)
    # name_idを取得
    name_id = a[0]

    name_id_row_index = return_row_index(symptom_i,name_id)
    print(name_id_row_index)
    # name_id_row_indexには　None　か数値が入る
    if name_id_row_index:
        # 行を挿入する。
        insert_index = name_id_row_index + 1
        # symptom_infoシートに登録
        # 行の挿入
        symptom_i.insert_rows(insert_index)
        for i in range(length):
            val = a[i]
            col = i + 1
            symptom_i.cell(row=insert_index, column=col).value = val
        wb_db.save(path_db)

    else:
        symptom_i.append(a)
        wb_db.save(path_db)

def output_excel():
    check = os.path.exists("./excel/teamdata.xlsx")
    # ファイルがあれば

make_dict_first_col_value(ws_players, players_info_dict)
make_dict_first_col_value(ws_symptom, symptom_info_dict)
make_dict_first_col_value(ws_name_table, name_table_dict)
make_dict_first_col_value(ws_body_table, body_table_dict)
make_dict_first_col_value(ws_symptom_table, symptom_table_dict)




class TopFrame(ttk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.create_widgets()

    def create_widgets(self):
        counts = 20
        for count_y in range(counts):
            self.empty_label = ttk.Label(self, text=count_y)
            self.empty_label.grid(row=count_y)
            for count_x in range(15):
                self.empty_label = ttk.Label(self, text=count_x)
                self.empty_label.grid(row=count_y, column=count_x)

        #label
        self.label_title = ttk.Label(self, text="Team Management")
        self.label_title.grid(row=6, column=7)
        #go to players
        self.button_register_players = ttk.Button(self, text="players", command=lambda: change_page(main_frame))
        self.button_register_players.grid(row=10, column=7)
        #go to information
        self.button_register_info = ttk.Button(self, text="information", command=lambda: change_page(info_frame))
        self.button_register_info.grid(row=10, column=8)

class Main(ttk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.create_widgets()

    def create_widgets(self):
        self.master.title(u"main_frame")

        counts = 20
        for count_y in range(counts):
            self.empty_label = ttk.Label(self, text=count_y)
            self.empty_label.grid(row=count_y)
            for count_x in range(15):
                self.empty_label = ttk.Label(self, text=count_x)
                self.empty_label.grid(row=count_y, column=count_x)

        #name part
        self.label_name = ttk.Label(self, text=players_information[1])
        self.entry_name = ttk.Entry(self)
        self.label_name.grid(row=5, column=0)
        self.entry_name.grid(row=5, column=1)
        #age part
        self.label_age = ttk.Label(self, text=players_information[2])
        self.entry_age = ttk.Entry(self)
        self.label_age.grid(row=6, column=0)
        self.entry_age.grid(row=6, column=1, pady=1)
        #sex part
        self.label_sex = ttk.Label(self, text=players_information[3])
        self.entry_sex = ttk.Entry(self)
        self.label_sex.grid(row=7, column=0)
        self.entry_sex.grid(row=7, column=1, pady=1)
        #position part
        self.label_position = ttk.Label(self, text=players_information[4])
        self.entry_position = ttk.Entry(self)
        self.label_position.grid(row=8, column=0)
        self.entry_position.grid(row=8, column=1, pady=1)

        self.button_back = ttk.Button(self, text=u"back", command= lambda:change_page(top_frame))
        self.button_back.grid(row=9, column=2, padx=1)
        self.button_register = ttk.Button(self,
                text=u"register",
                command= lambda:register_player(self.entry_name, self.entry_age, self.entry_sex, self.entry_position)
                )
        self.button_register.grid(row=9, column=3, padx=2)


class InfoFrame(ttk.Frame):
    def __init__(self, master=None):
        super().__init__(master)
        self.create_widgets()

    def create_widgets(self):
        self.master.title("info_frame")
        # id項目を取り出すための初期化
        self.names = []
        self.verletzungsarts = []
        self.korperteils = []

        ws_name_table = return_ws_db("name_table")
        ws_symptom_table = return_ws_db("symptom_table")
        ws_body_table = return_ws_db("body_table")

        # エクセルのシートから値を取得
        for cols in ws_name_table.iter_cols(min_col=2):
            for col in cols:
                self.names.append(col.value)

        for cols in ws_symptom_table.iter_cols(min_col=2):
            for col in cols:
                self.verletzungsarts.append(col.value)

        for cols in ws_body_table.iter_cols(min_col=2):
            for col in cols:
                self.korperteils.append(col.value)

        names_dict = id_value_key_dict(ws_name_table)

        counts = 20
        for count_y in range(counts):
            self.empty_label = ttk.Label(self, text=count_y)
            self.empty_label.grid(row=count_y)
            for count_x in range(15):
                self.empty_label = ttk.Label(self, text=count_x)
                self.empty_label.grid(row=count_y, column=count_x)

        # 表示させない空のentry
        self.entry_name_id = ttk.Entry(self)
        self.entry_verletzungsart_id = ttk.Entry(self)
        self.entry_korperteil_id = ttk.Entry(self)


        # playersname を表示
        self.combo_name = ttk.Combobox(self, state="readonly")
        self.combo_name["value"] = self.names
        self.combo_name.current(0)

        self.button_select = ttk.Button(self, text="select", command=lambda:show_select_info(self.combo_name.get(), names_dict))

        self.combo_name.grid(row=4, column=1, pady=2)
        self.button_select.grid(row=4, column=2, pady=2)

        # 項目一覧
        symptom_information = row_list(ws_symptom, min=0, max=1)
        print(symptom_information)
        #injury part
        self.label_verletzungsart = ttk.Label(self, text="Verletzungsart")
        # self.entry_verletzungsart = ttk.Entry(self)
        self.combo_verletzungsart = ttk.Combobox(self, state="readonly")
        self.combo_verletzungsart["value"] = self.verletzungsarts
        # self.combo_verletzungsart.current(0)

        self.label_verletzungsart.grid(row=5, column=0)
        self.combo_verletzungsart.grid(row=5, column=1)
        #part of injury part
        self.label_korperteil = ttk.Label(self, text="Korperteil")
        # self.entry_korperteil = ttk.Entry(self)
        self.combo_korperteil = ttk.Combobox(self, state="readonly")
        self.combo_korperteil["value"] = self.korperteils
        # self.combo_korperteil.current(0)


        self.label_korperteil.grid(row=6, column=0)
        self.combo_korperteil.grid(row=6, column=1, pady=1)
        #day of injury part
        self.label_verletzungsdatum = ttk.Label(self, text=symptom_information[0][3])
        self.entry_verletzungsdatum = ttk.Entry(self)
        self.label_verletzungsdatum.grid(row=7, column=0)
        self.entry_verletzungsdatum.grid(row=7, column=1, pady=1)
        #fully healed part
        self.label_der_geheiligte_tag = ttk.Label(self, text=symptom_information[0][4])
        self.entry_der_geheiligte_tag = ttk.Entry(self)
        self.label_der_geheiligte_tag.grid(row=8, column=0)
        self.entry_der_geheiligte_tag.grid(row=8, column=1, pady=1)
        #wie part
        self.label_wie = ttk.Label(self, text=symptom_information[0][5])
        self.entry_wie = ttk.Entry(self)
        self.label_wie.grid(row=9, column=0)
        self.entry_wie.grid(row=9, column=1, pady=1)
        #injuried place
        self.label_verletzter_ort = ttk.Label(self, text=symptom_information[0][6])
        self.entry_verletzter_ort = ttk.Entry(self)
        self.label_verletzter_ort.grid(row=10, column=0)
        self.entry_verletzter_ort.grid(row=10, column=1, pady=1)
        #diagnose
        self.label_diagnose = ttk.Label(self, text=symptom_information[0][7])
        self.entry_diagnose = ttk.Entry(self)
        self.label_diagnose.grid(row=11, column=0)
        self.entry_diagnose.grid(row=11, column=1, pady=1)
        #Verletzter Ort
        self.label_sonstig = ttk.Label(self, text=symptom_information[0][8])
        self.entry_sonstig = ttk.Entry(self)
        self.label_sonstig.grid(row=12, column=0)
        self.entry_sonstig.grid(row=12, column=1, pady=1)
        # 戻るボタン
        self.button_back = ttk.Button(self, text="back", command= lambda:change_page(top_frame))
        self.button_back.grid(row=12, column=2, padx=1)
        # 登録ボタン
        self.button_register = ttk.Button(
            self,
            text="register",
            command= lambda:register_symptom(
                    self.entry_name_id,
                    self.entry_verletzungsart_id,
                    self.entry_korperteil_id,
                    self.entry_verletzungsdatum,
                    self.entry_der_geheiligte_tag,
                    self.entry_wie,
                    self.entry_verletzter_ort,
                    self.entry_diagnose,
                    self.entry_sonstig,
                    )
            )
        self.button_register.grid(row=12, column=3, padx=2)
        # 一覧出力ボタン
        self.button_output = ttk.Button(self, text="output", command= lambda:output_excel())
        self.button_register.grid(row=11, column=2, padx=2)

if __name__ == "__main__":
    root = tk.Tk()
    root.geometry("800x600")
    main_frame = Main(root)
    top_frame = TopFrame(root)
    info_frame = InfoFrame(root)

    main_frame.grid(row=0, column=0, sticky=(tk.N,tk.W,tk.S,tk.E))
    top_frame.grid(row=0, column=0, sticky=(tk.N,tk.W,tk.S,tk.E))
    info_frame.grid(row=0, column=0, sticky=(tk.N,tk.W,tk.S,tk.E))

    top_frame.tkraise()

    root.mainloop()
